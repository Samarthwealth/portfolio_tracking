import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

def generate_excel_report(client_name, db, report_type, report_period):
    """Generate comprehensive Excel report"""
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb['Sheet'])  # Remove default sheet

        # Create summary sheet
        ws_summary = wb.create_sheet(title="Portfolio Summary")
        create_summary_sheet(ws_summary, client_name, db)

        # Create holdings sheet
        ws_holdings = wb.create_sheet(title="Current Holdings")
        create_holdings_sheet(ws_holdings, client_name, db)

        # Create transactions sheet
        ws_transactions = wb.create_sheet(title="All Transactions")
        create_transactions_sheet(ws_transactions, client_name, db)

        # Create cash movements sheet
        ws_cash = wb.create_sheet(title="Cash Movements")
        create_cash_movements_sheet(ws_cash, client_name, db)

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None

def create_summary_sheet(ws, client_name, db):
    """Create portfolio summary sheet"""

    # Title
    ws['A1'] = f"Portfolio Summary - {client_name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(color="FFFFFF", size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Report date
    ws['A2'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:D2')

    # Get data
    portfolio_summary = db.get_portfolio_summary(client_name)
    cash_balance = db.get_cash_balance(client_name)
    realized_profits = db.get_total_realized_profit(client_name)
    client_data = db.get_client_data(client_name)

    initial_cash = client_data[2] if client_data else 0

    def format_currency(amount):
        return f"₹{amount:,.2f}"

    # Key metrics
    row = 4
    metrics = [
        ("Initial Investment", format_currency(initial_cash)),
        ("Cash Balance", format_currency(cash_balance)),
        ("Invested Amount", format_currency(portfolio_summary.get('invested_amount', 0))),
        ("Current Portfolio Value", format_currency(portfolio_summary.get('current_value', 0))),
        ("Total Portfolio Value", format_currency(cash_balance + portfolio_summary.get('current_value', 0))),
        ("", ""),
        ("Unrealized P&L", format_currency(portfolio_summary.get('total_pnl', 0))),
        ("Realized P&L", format_currency(realized_profits)),
        ("Total P&L", format_currency(realized_profits + portfolio_summary.get('total_pnl', 0))),
    ]

    for metric, value in metrics:
        if metric:  # Skip empty rows
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Auto adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

def create_holdings_sheet(ws, client_name, db):
    """Create current holdings sheet"""

    # Title
    ws['A1'] = "Current Holdings"
    ws['A1'].font = Font(size=14, bold=True)

    # Get holdings data
    holdings_df = db.get_current_holdings_with_realized(client_name)

    if holdings_df.empty:
        ws['A3'] = "No current holdings found"
        return

    # Headers
    headers = [
        'Stock Symbol', 'Quantity', 'Avg Price', 'Current Price',
        'Current Value', 'Unrealized P&L', 'Unrealized %',
        'Realized P&L', 'Total P&L'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for idx, (_, row) in enumerate(holdings_df.iterrows(), 3):
        ws[f'A{idx}'] = row['stock_symbol']
        ws[f'B{idx}'] = row['quantity']
        ws[f'C{idx}'] = f"₹{row['avg_price']:,.2f}"
        ws[f'D{idx}'] = f"₹{row['current_price']:,.2f}"
        ws[f'E{idx}'] = f"₹{row['current_value']:,.2f}"
        ws[f'F{idx}'] = f"₹{row['unrealized_pnl']:,.2f}"
        ws[f'G{idx}'] = f"{row['unrealized_pnl_pct']:+.2f}%"
        ws[f'H{idx}'] = f"₹{row.get('realized_pnl', 0):,.2f}"
        ws[f'I{idx}'] = f"₹{(row['unrealized_pnl'] + row.get('realized_pnl', 0)):,.2f}"

    # Auto adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

def create_transactions_sheet(ws, client_name, db):
    """Create all transactions sheet"""

    # Title
    ws['A1'] = "All Transactions"
    ws['A1'].font = Font(size=14, bold=True)

    # Get transactions data
    transactions_df = db.get_all_transactions_with_realized(client_name)

    if transactions_df.empty:
        ws['A3'] = "No transactions found"
        return

    # Headers
    headers = [
        'Date', 'Stock Symbol', 'Type', 'Quantity', 'Price',
        'Total Amount', 'Brokerage', 'Realized P&L'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for idx, (_, row) in enumerate(transactions_df.iterrows(), 3):
        ws[f'A{idx}'] = row['date']
        ws[f'B{idx}'] = row['stock_symbol']
        ws[f'C{idx}'] = row['transaction_type']
        ws[f'D{idx}'] = row['quantity']
        ws[f'E{idx}'] = f"₹{row['price']:,.2f}"
        ws[f'F{idx}'] = f"₹{row['total_amount']:,.2f}"
        ws[f'G{idx}'] = f"₹{row['brokerage']:,.2f}"
        ws[f'H{idx}'] = f"₹{row.get('realized_profit', 0):,.2f}"

    # Auto adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

def create_cash_movements_sheet(ws, client_name, db):
    """Create cash movements sheet"""

    # Title
    ws['A1'] = "Cash Movements"
    ws['A1'].font = Font(size=14, bold=True)

    # Get cash transactions
    cash_df = db.get_cash_transactions(client_name)

    if cash_df.empty:
        ws['A3'] = "No cash transactions found"
        return

    # Headers
    headers = ['Date', 'Type', 'Amount', 'Description', 'Running Balance']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Calculate running balance and add data rows
    running_balance = 0
    for idx, (_, row) in enumerate(cash_df.iterrows(), 3):
        if row['transaction_type'] == 'Deposit':
            running_balance += row['amount']
        else:
            running_balance -= row['amount']

        ws[f'A{idx}'] = row['date']
        ws[f'B{idx}'] = row['transaction_type']
        ws[f'C{idx}'] = f"₹{row['amount']:,.2f}"
        ws[f'D{idx}'] = row.get('description', '')
        ws[f'E{idx}'] = f"₹{running_balance:,.2f}"

    # Auto adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
